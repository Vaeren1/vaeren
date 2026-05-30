"""Strukturelle Frage-Extraktion aus xlsx (Spec §6.2, Tier 1).

Heuristik zur Header-Erkennung:
1. Suche in den ersten Zeilen nach einer Header-Zeile, deren Zellen ein
   Frage-Schlüsselwort (``frage``/``question``) UND idealerweise ein
   Antwort-Schlüsselwort (``antwort``/``answer``) enthalten.
2. Frage-Spalte = Spalte mit Frage-Schlüsselwort; Antwort-Spalte = Spalte mit
   Antwort-Schlüsselwort, sonst die direkt rechts neben der Frage-Spalte;
   Nummer-Spalte = Spalte mit ``nr``/``no``/``#`` (optional).
3. Fallback ohne Header: erste Spalte mit Text = Frage, nächste Spalte = Antwort.

Pro Datenzeile entsteht ``{nummer, text, feld_referenz: {sheet, antwort_zelle}}``.
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter

_FRAGE_KEYS = ("frage", "question")
_ANTWORT_KEYS = ("antwort", "answer", "kommentar", "comment", "response")
_NUMMER_KEYS = ("nr", "no", "#", "nummer", "number", "id")
_HEADER_SUCHTIEFE = 10  # nur die obersten Zeilen nach einem Header durchsuchen


def extrahiere_fragen_xlsx(pfad: str) -> list[dict]:
    """Extrahiert Fragen + Antwort-Zellreferenzen aus allen Sheets der Datei."""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(pfad, data_only=True, read_only=True)
    except Exception:
        return []

    fragen: list[dict] = []
    try:
        for ws in wb.worksheets:
            fragen.extend(_extrahiere_sheet(ws))
    finally:
        wb.close()
    return fragen


def _extrahiere_sheet(ws) -> list[dict]:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        return []

    header = _finde_header(rows)
    if header is not None:
        return _extrahiere_mit_header(ws.title, rows, *header)
    return _extrahiere_fallback(ws.title, rows)


def _finde_header(rows: list[list]) -> tuple[int, int, int, int | None] | None:
    """Liefert ``(header_zeile_idx, frage_col, antwort_col, nummer_col)`` oder None."""
    for idx, row in enumerate(rows[:_HEADER_SUCHTIEFE]):
        frage_col = _spalte_mit_key(row, _FRAGE_KEYS)
        if frage_col is None:
            continue
        antwort_col = _spalte_mit_key(row, _ANTWORT_KEYS)
        if antwort_col is None:
            antwort_col = frage_col + 1
        nummer_col = _spalte_mit_key(row, _NUMMER_KEYS)
        return idx, frage_col, antwort_col, nummer_col
    return None


def _spalte_mit_key(row: list, keys: tuple[str, ...]) -> int | None:
    for col, zelle in enumerate(row):
        if zelle is None:
            continue
        text = str(zelle).strip().lower()
        if any(k in text for k in keys):
            return col
    return None


def _extrahiere_mit_header(
    sheet: str,
    rows: list[list],
    header_idx: int,
    frage_col: int,
    antwort_col: int,
    nummer_col: int | None,
) -> list[dict]:
    fragen: list[dict] = []
    for offset, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        text = _zell_text(row, frage_col)
        if not text:
            continue
        nummer = _zell_text(row, nummer_col) if nummer_col is not None else None
        fragen.append(
            _baue_frage(sheet, nummer, text, zeile=offset, antwort_col=antwort_col)
        )
    return fragen


def _extrahiere_fallback(sheet: str, rows: list[list]) -> list[dict]:
    """Ohne Header: erste Text-Spalte = Frage, Spalte direkt rechts = Antwort."""
    frage_col = _erste_textspalte(rows)
    if frage_col is None:
        return []
    antwort_col = frage_col + 1
    fragen: list[dict] = []
    for offset, row in enumerate(rows, start=1):
        text = _zell_text(row, frage_col)
        if not text:
            continue
        fragen.append(
            _baue_frage(sheet, None, text, zeile=offset, antwort_col=antwort_col)
        )
    return fragen


def _erste_textspalte(rows: list[list]) -> int | None:
    max_cols = max((len(r) for r in rows), default=0)
    for col in range(max_cols):
        if any(_zell_text(row, col) for row in rows):
            return col
    return None


def _baue_frage(
    sheet: str, nummer: str | None, text: str, *, zeile: int, antwort_col: int
) -> dict:
    antwort_zelle = f"{get_column_letter(antwort_col + 1)}{zeile}"
    return {
        "nummer": nummer,
        "text": text,
        "feld_referenz": {"sheet": sheet, "antwort_zelle": antwort_zelle},
    }


def _zell_text(row: list, col: int | None) -> str:
    if col is None or col >= len(row):
        return ""
    wert = row[col]
    return str(wert).strip() if wert is not None else ""
