"""API-Flow-, Permission- + RDG-Gate-Tests für den OEM-Fragebogen-Auswerter
(Feature 4, Phase F3).

Voller Flow (LLM gemockt, kleine xlsx-Fixture):
  upload → vorschlagen → seite bestätigen → final-attestieren → export(xlsx).

Kritische Gates:
- RDG: export VOR final-attestieren → 409.
- Bibliothek-Übernahme bei finaler Attestierung verifiziert.
- Permission: Mitarbeiter → 403, GF + Compliance → ok.
"""

from __future__ import annotations

import io

import openpyxl
import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from django_tenants.utils import schema_context


@pytest.fixture
def _media_tmp(settings, tmp_path):
    """FileField-Schreibziel auf tmp_path lenken (sonst landet es im cwd)."""
    settings.MEDIA_ROOT = str(tmp_path)
    return tmp_path


@pytest.fixture
def _xlsx_upload():
    """Kleine xlsx-Fixture mit Header Nr|Frage|Antwort + zwei Fragen."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Nr"
    ws["B1"] = "Frage"
    ws["C1"] = "Antwort"
    ws["A2"] = "1"
    ws["B2"] = "Haben Sie ein ISMS?"
    ws["A3"] = "2"
    ws["B3"] = "Führen Sie ein Datenpannen-Register?"
    buf = io.BytesIO()
    wb.save(buf)
    return SimpleUploadedFile(
        "vda.xlsx",
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@pytest.fixture
def _pdf_unstrukturiert_upload():
    """Minimales Text-PDF ohne AcroForm → wird als pdf_unstrukturiert (Tier 2) erkannt."""
    from reportlab.pdfgen import canvas

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(595, 842))
    c.drawString(72, 800, "Haben Sie ein ISMS? Antwort:")
    c.showPage()
    c.save()
    return SimpleUploadedFile(
        "oem-scan.pdf", buf.getvalue(), content_type="application/pdf"
    )


def _mock_engine(monkeypatch):
    """Patcht die LLM-Grenze der Antwort-Engine (RDG-valider Vorschlagstext)."""
    monkeypatch.setattr(
        "fragebogen.answer_engine._llm_antwort",
        lambda frage, snippets: {
            "text": "Nach unserer Einschätzung: Ja, vorhanden. Bitte prüfen.",
            "quellen_referenzen": [],
            "confidence": 0.8,
        },
    )
    # Evidenz-Pool ohne DB-Abhängigkeit (Aggregatoren brauchen aktive Module).
    monkeypatch.setattr("fragebogen.views.sammle_evidenz", lambda: [])


@pytest.mark.django_db
def test_voller_flow_xlsx(tenant_client_gf, _xlsx_upload, _media_tmp, monkeypatch):
    _mock_engine(monkeypatch)
    c = tenant_client_gf

    # upload
    r = c.post(
        "/api/fragebogen/upload/",
        {"datei": _xlsx_upload, "quelle_oem": "VDA ISA"},
    )
    assert r.status_code == 201, r.content
    fb = r.json()
    fb_id = fb["id"]
    assert fb["format"] == "xlsx"
    assert fb["tier"] == 1
    assert fb["status"] == "analysiert"
    assert len(fb["fragen"]) == 2

    # vorschlagen
    r = c.post(f"/api/fragebogen/{fb_id}/vorschlagen/")
    assert r.status_code == 200, r.content
    data = r.json()
    assert data["status"] == "vorgeschlagen"
    assert all(f["antwort"]["entwurf_text"] for f in data["fragen"])
    assert all(f["antwort"]["confidence"] == pytest.approx(0.8) for f in data["fragen"])

    # seiten (Struktur-Daten, eine Seite)
    r = c.get(f"/api/fragebogen/{fb_id}/seiten/")
    assert r.status_code == 200, r.content
    seiten = r.json()["seiten"]
    assert len(seiten) == 1
    assert seiten[0]["seite"] == 1
    assert seiten[0]["bestaetigt"] is False

    # Export VOR Attestierung → 409 (RDG-Gate)
    r = c.post(f"/api/fragebogen/{fb_id}/export/")
    assert r.status_code == 409, r.content

    # seite bestätigen
    r = c.post(f"/api/fragebogen/{fb_id}/seite/1/bestaetigen/")
    assert r.status_code == 200, r.content
    assert 1 in r.json()["bestaetigte_seiten"]

    # final-attestieren
    r = c.post(f"/api/fragebogen/{fb_id}/final-attestieren/")
    assert r.status_code == 200, r.content
    assert r.json()["final_attestiert"] is True

    # export (xlsx) → befüllte Datei + Status exportiert
    r = c.post(f"/api/fragebogen/{fb_id}/export/")
    assert r.status_code == 200, r.content
    assert r.json()["status"] == "exportiert"
    assert r.json()["export_datei_url"]

    # export-status
    r = c.get(f"/api/fragebogen/{fb_id}/export-status/")
    assert r.status_code == 200, r.content
    assert r.json()["export_bereit"] is True

    # Verifizieren: befüllte Datei enthält die Antwort in der Antwort-Zelle.
    from fragebogen.models import Fragebogen

    with schema_context(tenant_client_gf_schema(tenant_client_gf)):
        fb_obj = Fragebogen.objects.get(pk=fb_id)
        wb = openpyxl.load_workbook(fb_obj.export_datei.path)
        ws = wb.active
        assert "Nach unserer Einschätzung" in str(ws["C2"].value)
        assert ws["B2"].value == "Haben Sie ein ISMS?"  # Original unverändert


@pytest.mark.django_db
def test_upload_pdf_unstrukturiert_extrahiert_via_ocr(
    tenant_client_gf, _pdf_unstrukturiert_upload, _media_tmp, monkeypatch
):
    """C1: Ein unstrukturiertes PDF wird als Tier 2 erkannt und der OCR-Extraktor
    (`extrahiere_fragen_ocr`) verdrahtet — Upload erzeugt Fragen statt 0."""
    # OCR-Grenze mocken (kein echter pdf2image/tesseract/LLM-Call) → 2 Fragen.
    # Der Extraktor wird über _EXTRAKTOREN (Dict) aufgelöst — daher den Dict-
    # Eintrag patchen, nicht nur das Modul-Attribut.
    from fragebogen import views as fb_views

    fake_ocr = lambda pfad: [  # noqa: E731
        {
            "text": "Haben Sie ein ISMS?",
            "feld_referenz": {"seite": 1, "bbox": [10, 20, 30, 40], "schrift_pt": 11},
            "extraktion_quelle": "ocr",
        },
        {
            "text": "Führen Sie ein Datenpannen-Register?",
            "feld_referenz": {"seite": 1, "bbox": [10, 60, 30, 40], "schrift_pt": 11},
            "extraktion_quelle": "ocr",
        },
    ]
    monkeypatch.setitem(fb_views._EXTRAKTOREN, "pdf_unstrukturiert", fake_ocr)

    r = tenant_client_gf.post(
        "/api/fragebogen/upload/", {"datei": _pdf_unstrukturiert_upload}
    )
    assert r.status_code == 201, r.content
    fb = r.json()
    assert fb["format"] == "pdf_unstrukturiert"
    assert fb["tier"] == 2
    assert len(fb["fragen"]) == 2
    assert {f["text"] for f in fb["fragen"]} == {
        "Haben Sie ein ISMS?",
        "Führen Sie ein Datenpannen-Register?",
    }


def tenant_client_gf_schema(client):
    """Liest das Tenant-Schema aus dem Test-Client-Host (HTTP_HOST → Tenant)."""
    from django.db import connection

    from tenants.models import TenantDomain

    connection.set_schema_to_public()
    with schema_context("public"):
        host = client.defaults["HTTP_HOST"]
        dom = TenantDomain.objects.get(domain=host)
        return dom.tenant.schema_name


@pytest.mark.django_db
def test_export_vor_attestierung_409(tenant_client_gf, _xlsx_upload, _media_tmp, monkeypatch):
    _mock_engine(monkeypatch)
    c = tenant_client_gf
    r = c.post("/api/fragebogen/upload/", {"datei": _xlsx_upload})
    fb_id = r.json()["id"]
    r = c.post(f"/api/fragebogen/{fb_id}/export/")
    assert r.status_code == 409, r.content


@pytest.mark.django_db
def test_bibliothek_uebernahme_bei_attestierung(
    tenant_client_gf, _xlsx_upload, _media_tmp, monkeypatch
):
    _mock_engine(monkeypatch)
    c = tenant_client_gf
    schema = tenant_client_gf_schema(c)

    r = c.post("/api/fragebogen/upload/", {"datei": _xlsx_upload})
    fb_id = r.json()["id"]
    c.post(f"/api/fragebogen/{fb_id}/vorschlagen/")
    c.post(f"/api/fragebogen/{fb_id}/seite/1/bestaetigen/")

    from fragebogen.models import AntwortBibliothekEintrag

    with schema_context(schema):
        assert AntwortBibliothekEintrag.objects.count() == 0

    r = c.post(f"/api/fragebogen/{fb_id}/final-attestieren/")
    assert r.status_code == 200, r.content

    with schema_context(schema):
        # zwei (unterschiedliche) Fragen → zwei Bibliothek-Einträge
        assert AntwortBibliothekEintrag.objects.count() == 2
        texte = {e.frage_kanonisch for e in AntwortBibliothekEintrag.objects.all()}
        assert "Haben Sie ein ISMS?" in texte


@pytest.mark.django_db
def test_verwendungs_count_steigt(tenant_client_gf, _xlsx_upload, _media_tmp, monkeypatch):
    _mock_engine(monkeypatch)
    c = tenant_client_gf
    schema = tenant_client_gf_schema(c)

    # Bibliothek vorab mit einem zur ersten Frage passenden Eintrag füllen.
    from fragebogen.models import AntwortBibliothekEintrag

    with schema_context(schema):
        AntwortBibliothekEintrag.objects.create(
            frage_kanonisch="Haben Sie ein ISMS etabliert?",
            antwort_text="Ja, seit 2025.",
        )

    r = c.post("/api/fragebogen/upload/", {"datei": _xlsx_upload})
    fb_id = r.json()["id"]
    c.post(f"/api/fragebogen/{fb_id}/vorschlagen/")

    with schema_context(schema):
        eintrag = AntwortBibliothekEintrag.objects.get(
            frage_kanonisch="Haben Sie ein ISMS etabliert?"
        )
        assert eintrag.verwendungs_count == 1
        assert eintrag.zuletzt_verwendet is not None


@pytest.mark.django_db
def test_antwort_patch_setzt_editiert(
    tenant_client_gf, _xlsx_upload, _media_tmp, monkeypatch
):
    _mock_engine(monkeypatch)
    c = tenant_client_gf
    r = c.post("/api/fragebogen/upload/", {"datei": _xlsx_upload})
    fb_id = r.json()["id"]
    detail = c.post(f"/api/fragebogen/{fb_id}/vorschlagen/").json()
    aid = detail["fragen"][0]["antwort"]["id"]

    r = c.patch(
        f"/api/fragebogen/{fb_id}/antwort/{aid}/",
        {"bestaetigt_text": "Manuell editierte Antwort."},
        content_type="application/json",
    )
    assert r.status_code == 200, r.content
    frage0 = next(f for f in r.json()["fragen"] if f["antwort"]["id"] == aid)
    assert frage0["antwort"]["status"] == "editiert"
    assert frage0["antwort"]["finaler_text"] == "Manuell editierte Antwort."


@pytest.mark.django_db
def test_final_attestieren_ohne_alle_seiten_409(
    tenant_client_gf, _xlsx_upload, _media_tmp, monkeypatch
):
    _mock_engine(monkeypatch)
    c = tenant_client_gf
    r = c.post("/api/fragebogen/upload/", {"datei": _xlsx_upload})
    fb_id = r.json()["id"]
    c.post(f"/api/fragebogen/{fb_id}/vorschlagen/")
    # keine Seite bestätigt
    r = c.post(f"/api/fragebogen/{fb_id}/final-attestieren/")
    assert r.status_code == 409, r.content
    assert 1 in r.json()["offene_seiten"]


# --- RDG-Layer-2-Gate (rdg_ok) --------------------------------------------


def _mock_engine_rdg_verstoss(monkeypatch):
    """Patcht die LLM-Grenze mit einem Text, der RDG-Layer-2 verletzt."""
    monkeypatch.setattr(
        "fragebogen.answer_engine._llm_antwort",
        lambda frage, snippets: {
            # 'rechtlich verpflichtet' ist eine verbotene Pflichtformulierung.
            "text": "Sie sind rechtlich verpflichtet, ein ISMS zu betreiben.",
            "quellen_referenzen": [],
            "confidence": 0.8,
        },
    )
    monkeypatch.setattr("fragebogen.views.sammle_evidenz", lambda: [])


@pytest.mark.django_db
def test_rdg_verstoss_nicht_exportiert_nicht_uebernommen(
    tenant_client_gf, _xlsx_upload, _media_tmp, monkeypatch
):
    """rdg_ok=False + status ENTWURF → NICHT in Bibliothek + Export leer."""
    _mock_engine_rdg_verstoss(monkeypatch)
    c = tenant_client_gf
    schema = tenant_client_gf_schema(c)

    r = c.post("/api/fragebogen/upload/", {"datei": _xlsx_upload})
    fb_id = r.json()["id"]

    detail = c.post(f"/api/fragebogen/{fb_id}/vorschlagen/").json()
    # Engine hat RDG-Verstoß erkannt → rdg_ok=False auf allen Antworten.
    assert all(f["antwort"]["rdg_ok"] is False for f in detail["fragen"])

    # seiten-Action gibt rdg_ok ebenfalls aus.
    seiten = c.get(f"/api/fragebogen/{fb_id}/seiten/").json()["seiten"]
    assert seiten[0]["fragen"][0]["rdg_ok"] is False

    c.post(f"/api/fragebogen/{fb_id}/seite/1/bestaetigen/")
    r = c.post(f"/api/fragebogen/{fb_id}/final-attestieren/")
    assert r.status_code == 200, r.content

    from fragebogen.models import AntwortBibliothekEintrag

    with schema_context(schema):
        # Verbotene Formulierung darf NICHT als Wissen recycelt werden.
        assert AntwortBibliothekEintrag.objects.count() == 0

    # Export: die rdg_ok=False-Antworten bleiben leer (Zelle nicht befüllt).
    r = c.post(f"/api/fragebogen/{fb_id}/export/")
    assert r.status_code == 200, r.content

    from fragebogen.models import Fragebogen

    with schema_context(schema):
        fb_obj = Fragebogen.objects.get(pk=fb_id)
        wb = openpyxl.load_workbook(fb_obj.export_datei.path)
        ws = wb.active
        # C2 (Antwort-Zelle der ersten Frage) bleibt leer.
        assert ws["C2"].value in (None, "")


@pytest.mark.django_db
def test_rdg_verstoss_nach_editieren_normal_exportiert(
    tenant_client_gf, _xlsx_upload, _media_tmp, monkeypatch
):
    """rdg_ok=False, aber nach PATCH (status EDITIERT) → normal exportiert + übernommen."""
    _mock_engine_rdg_verstoss(monkeypatch)
    c = tenant_client_gf
    schema = tenant_client_gf_schema(c)

    r = c.post("/api/fragebogen/upload/", {"datei": _xlsx_upload})
    fb_id = r.json()["id"]
    detail = c.post(f"/api/fragebogen/{fb_id}/vorschlagen/").json()

    # Beide Antworten manuell umformulieren (Mensch greift ein → status EDITIERT).
    for f in detail["fragen"]:
        aid = f["antwort"]["id"]
        r = c.patch(
            f"/api/fragebogen/{fb_id}/antwort/{aid}/",
            {"bestaetigt_text": "Nach unserer Einschätzung vorhanden. Bitte prüfen."},
            content_type="application/json",
        )
        assert r.status_code == 200, r.content

    c.post(f"/api/fragebogen/{fb_id}/seite/1/bestaetigen/")
    r = c.post(f"/api/fragebogen/{fb_id}/final-attestieren/")
    assert r.status_code == 200, r.content

    from fragebogen.models import AntwortBibliothekEintrag, Fragebogen

    with schema_context(schema):
        # Editierte Antworten gehen normal in die Bibliothek.
        assert AntwortBibliothekEintrag.objects.count() == 2

    r = c.post(f"/api/fragebogen/{fb_id}/export/")
    assert r.status_code == 200, r.content

    with schema_context(schema):
        fb_obj = Fragebogen.objects.get(pk=fb_id)
        wb = openpyxl.load_workbook(fb_obj.export_datei.path)
        ws = wb.active
        assert "Nach unserer Einschätzung" in str(ws["C2"].value)


# --- Upload-Sicherheit (Whitelist + Größe) --------------------------------


@pytest.mark.django_db
def test_upload_exe_abgelehnt(tenant_client_gf, _media_tmp):
    bad = SimpleUploadedFile(
        "malware.exe", b"MZ\x90\x00", content_type="application/octet-stream"
    )
    r = tenant_client_gf.post("/api/fragebogen/upload/", {"datei": bad})
    assert r.status_code == 400, r.content
    assert "nicht erlaubt" in r.json()["detail"].lower()


@pytest.mark.django_db
def test_upload_zu_gross_abgelehnt(tenant_client_gf, _media_tmp):
    # 26 MB > 25-MB-Limit. Endung .xlsx ist erlaubt — Größe muss greifen.
    big = SimpleUploadedFile(
        "gross.xlsx", b"\x00" * (26 * 1024 * 1024), content_type="application/octet-stream"
    )
    r = tenant_client_gf.post("/api/fragebogen/upload/", {"datei": big})
    assert r.status_code == 400, r.content
    assert "zu groß" in r.json()["detail"].lower()


# --- Seiten-Nr-Validierung -------------------------------------------------


@pytest.mark.django_db
def test_unbekannte_seite_bestaetigen_404(
    tenant_client_gf, _xlsx_upload, _media_tmp, monkeypatch
):
    _mock_engine(monkeypatch)
    c = tenant_client_gf
    r = c.post("/api/fragebogen/upload/", {"datei": _xlsx_upload})
    fb_id = r.json()["id"]
    # Fixture hat nur Seite 1 → Seite 99 existiert nicht.
    r = c.post(f"/api/fragebogen/{fb_id}/seite/99/bestaetigen/")
    assert r.status_code == 404, r.content


# --- Permissions ----------------------------------------------------------


@pytest.mark.django_db
def test_compliance_darf(tenant_client_compliance, _xlsx_upload, _media_tmp, monkeypatch):
    _mock_engine(monkeypatch)
    r = tenant_client_compliance.post(
        "/api/fragebogen/upload/", {"datei": _xlsx_upload}
    )
    assert r.status_code == 201, r.content


@pytest.mark.django_db
def test_mitarbeiter_darf_nicht(tenant_client_mitarbeiter):
    r = tenant_client_mitarbeiter.get("/api/fragebogen/")
    assert r.status_code == 403, r.content


@pytest.mark.django_db
def test_mitarbeiter_darf_nicht_uploaden(tenant_client_mitarbeiter, _xlsx_upload):
    r = tenant_client_mitarbeiter.post(
        "/api/fragebogen/upload/", {"datei": _xlsx_upload}
    )
    assert r.status_code == 403, r.content


@pytest.mark.django_db
def test_mitarbeiter_darf_bibliothek_nicht(tenant_client_mitarbeiter):
    r = tenant_client_mitarbeiter.get("/api/antwort-bibliothek/")
    assert r.status_code == 403, r.content


@pytest.mark.django_db
def test_bibliothek_crud_gf(tenant_client_gf):
    c = tenant_client_gf
    r = c.post(
        "/api/antwort-bibliothek/",
        {"frage_kanonisch": "Test?", "antwort_text": "Ja."},
        content_type="application/json",
    )
    assert r.status_code == 201, r.content
    eintrag_id = r.json()["id"]

    r = c.get("/api/antwort-bibliothek/")
    assert r.status_code == 200
    assert any(e["id"] == eintrag_id for e in r.json()["results"])

    r = c.patch(
        f"/api/antwort-bibliothek/{eintrag_id}/",
        {"antwort_text": "Ja, geändert."},
        content_type="application/json",
    )
    assert r.status_code == 200, r.content
    assert r.json()["antwort_text"] == "Ja, geändert."

    r = c.delete(f"/api/antwort-bibliothek/{eintrag_id}/")
    assert r.status_code == 204, r.content
