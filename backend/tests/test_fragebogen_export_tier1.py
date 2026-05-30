"""Tier-1-Export-Tests (Phase E1 + PDF-Formular-Roundtrip).

Roundtrip xlsx: Antwort-Zelle gefüllt, Frage-Zelle unverändert.
Roundtrip PDF-AcroForm: Feldwert gesetzt, via pypdf get_fields() verifiziert.

Fixtures werden in tmp_path erzeugt (kein committetes Binär-Asset).
Workbooks werden mit wb.close() geschlossen (conftest: ResourceWarning -> error).
"""

from __future__ import annotations

import openpyxl

from fragebogen.export.fill_pdfform import fuelle_pdfform
from fragebogen.export.fill_xlsx import fuelle_xlsx


def test_xlsx_roundtrip(tmp_path):
    src = tmp_path / "vda.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B2"] = "Haben Sie ein ISMS?"
    wb.save(src)
    wb.close()

    out = tmp_path / "out.xlsx"
    fuelle_xlsx(str(src), str(out), {"C2": "Ja, ISMS seit 2025."})

    wb2 = openpyxl.load_workbook(out)
    assert wb2.active["C2"].value == "Ja, ISMS seit 2025."
    assert wb2.active["B2"].value == "Haben Sie ein ISMS?"  # Original unverändert
    wb2.close()


def test_xlsx_roundtrip_named_sheet(tmp_path):
    """Zelle mit Sheet-Prefix 'Tabelle1!C2' wird im richtigen Sheet gesetzt."""
    src = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fragen"
    ws["B2"] = "Frage?"
    wb.save(src)
    wb.close()

    out = tmp_path / "multi_out.xlsx"
    fuelle_xlsx(str(src), str(out), {"Fragen!C2": "Antwort."})

    wb2 = openpyxl.load_workbook(out)
    assert wb2["Fragen"]["C2"].value == "Antwort."
    wb2.close()


def test_xlsx_leere_antwort_uebersprungen(tmp_path):
    """Leere/None-Antworten lassen die Zelle unangetastet."""
    src = tmp_path / "x.xlsx"
    wb = openpyxl.Workbook()
    wb.active["C2"] = "alt"
    wb.save(src)
    wb.close()

    out = tmp_path / "x_out.xlsx"
    fuelle_xlsx(str(src), str(out), {"C2": ""})

    wb2 = openpyxl.load_workbook(out)
    assert wb2.active["C2"].value == "alt"
    wb2.close()


def _erzeuge_fillable_pdf(tmp_path, feldname: str):
    from reportlab.pdfgen import canvas

    p = tmp_path / "form.pdf"
    c = canvas.Canvas(str(p))
    c.drawString(100, 750, "Frage 1")
    c.acroForm.textfield(
        name=feldname,
        x=100,
        y=700,
        width=200,
        height=20,
        borderStyle="inset",
    )
    c.save()
    return p


def test_pdfform_roundtrip(tmp_path):
    from pypdf import PdfReader

    src = _erzeuge_fillable_pdf(tmp_path, feldname="antwort_1")
    out = tmp_path / "form_out.pdf"

    fuelle_pdfform(str(src), str(out), {"antwort_1": "Ja, seit 2025."})

    felder = PdfReader(str(out)).get_fields()
    assert felder["antwort_1"]["/V"] == "Ja, seit 2025."


def test_pdfform_unbekanntes_feld_ignoriert(tmp_path):
    """Ein nicht existierendes Feld im Mapping darf nicht crashen."""
    from pypdf import PdfReader

    src = _erzeuge_fillable_pdf(tmp_path, feldname="antwort_1")
    out = tmp_path / "form_out2.pdf"

    fuelle_pdfform(
        str(src),
        str(out),
        {"antwort_1": "Befuellt.", "gibt_es_nicht": "egal"},
    )

    felder = PdfReader(str(out)).get_fields()
    assert felder["antwort_1"]["/V"] == "Befuellt."
